import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

#same initial code as used previously to get extrapolated date
workbook = load_workbook(filename="Nat_Gas.xlsx")
ws = workbook.active

#function that takes a integer inputs for dates in the format MM,DD,YYYY and outputs a price
def natGasPrice(find, day):
    lower = allprices[find]
    upper = allprices[find + 1]
    return np.interp(day, [0, 30.5], [lower, upper])#30.5 average days per month

#function to make code less messy but still messy
def findDate(month, day, year):
    find = 0
    for i in range(year - 2021):
        find += 12
    for i in range(month):
        find += 1
    if year == 2020:
        if month == 10:
            return 0
        elif month == 11:
            return 1
        else:
            return 2
    return find + 2

#how many barrels to purchase
MMBtu = int(input("How MMBtu of natrual gas do you want to buy?: "))

#injection date
while True:
    print("\nEnter an injection date between the dates 10/31/2020 and 10/31/2025 for an estimate of the natrual gas price on that day")
    yearinj = int(input("Enter year: "))
    monthinj = int(input("Enter a month: "))
    dayinj = int(input("Enter a day: "))
    if yearinj == 2020: #stop invalid input when year is 2020 or 2025
        if monthinj < 10 or monthinj > 12 or dayinj < 1 or dayinj > 31:
            print("\nInput not valid\n")
    elif yearinj == 2025:
        if monthinj > 10 or monthinj < 1 or dayinj < 1 or dayinj > 31:
            print("\nInput not valid\n")
    elif yearinj > 2025 or yearinj < 2020 or monthinj < 1 or monthinj > 12 or dayinj > 31 or dayinj < 1: #stop invalid input
            print("\nInput not valid\n")
    else:
        break

#withdrawl date makes sure it is after chosen injection date
while True:
    print("\nEnter a withdrawl date between the dates " + str(monthinj) + "/" + str(dayinj) + "/" + str(yearinj) + " and 10/31/2025 for an estimate of the natrual gas price on that day")
    yearwith = int(input("Enter year: "))
    monthwith = int(input("Enter a month: "))
    daywith = int(input("Enter a day: "))
    if yearwith == 2020: #stop invalid input when year is 2020 or 2025
        if monthwith < 10 or monthwith > 12 or daywith < 1 or daywith > 31:
            print("\nInput not valid\n")
    elif yearwith == 2025:
        if monthwith > 10 or monthwith < 1 or daywith < 1 or daywith > 31:
            print("\nInput not valid\n")
    elif yearwith > 2025 or yearwith < 2020 or monthwith < 1 or monthwith > 12 or daywith > 31 or daywith < 1: #stop invalid input
            print("\nInput not valid\n")
    elif findDate(monthwith, daywith, yearwith) < findDate(monthinj, dayinj, yearinj):
        print("\nEnter a withdrawl date after the injection date\n")
    else:
        break

dates = []
prices = []

for col in ws['A']:
    dates.append(col.value)
dates.remove('Dates')

for col in ws['B']:
    prices.append(col.value)
prices.remove('Prices')

col = 0
row = 0
larray = np.zeros((12,4))#12 months 4 years of data

#sorting data into months
for int, price in enumerate(prices):
    if row == 12:
        row = 0
        col += 1
    larray[row][col] = price
    row += 1

#finding average rise each year in specific months
avgmonth = np.zeros((12,3))
avg = []
for i in range(12):
    for j in range(3):
        avgmonth[i][j] = larray[i][3-j] - larray[i][2-j]

#avg change per month over the 4 years
extprices = prices[-12:]
for i in range(12):
    extprices[i] += (avgmonth[i][0]+avgmonth[i][1]+avgmonth[i][2])/3

MMBtuPrice = MMBtu * 10**3
allprices = prices + extprices
foundDateInj = findDate(monthinj, dayinj, yearinj)
foundDateWith = findDate(monthwith, daywith, yearwith)
injPrice = round(natGasPrice(foundDateInj, dayinj), 2)#injection and withdrawl prices
withPrice = round(natGasPrice(foundDateWith, daywith), 2)
injPriceTot = round(injPrice * MMBtuPrice, 3)
withPriceTot = round(withPrice * MMBtuPrice, 3)
extrationPrice = MMBtu * 10 * 2#injection and withdrawl
storageTime = foundDateWith - foundDateInj + 1
storageTimePrice = storageTime * 10

print("\nThe price per Btu that this can be purchased at on " + str(monthinj) + "/" + str(dayinj) + "/" + str(yearinj) + " is $" + str(injPrice))
print("\nThe price per Btu that this can be sold at on " + str(monthwith) + "/" + str(daywith) + "/" + str(yearwith) + " is $" + str(withPrice))
print("\nYour purchase price is $" + str(injPriceTot) + "k and your sell price is $" + str(withPriceTot) + "k")
print("\nThe rate to inject and withdrawl natrual gas is $10k per MMBtu, therefore it will cost $" + str(extrationPrice) + "k")
print("\nThe commodities will need to be stored for " + str(storageTime) + " months, this will cost $" + str(storageTimePrice) + "k , that costing $10k per month")
print("\nTransportation will cost $100k to and from the storage facility" )

totalProfit = round(-injPriceTot + withPriceTot - storageTimePrice - 100 - extrationPrice, 3)
print("\nYour total profit is $" + str(totalProfit) + "k")